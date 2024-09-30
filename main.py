import time
from datetime import datetime
from dateutil.relativedelta import relativedelta
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import smtplib
import email
from email import encoders
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
import os

#Получить необходимые даты
current_date = datetime.now()
startDay = (current_date.replace(day=1) - relativedelta(months=1)).strftime("%d.%m.%Y")
lastDay = (current_date.replace(day=1) - relativedelta(days=1)).strftime("%d.%m.%Y")
start_day_num = (current_date.replace(day=1) - relativedelta(months=1)).day
start_month_num = (current_date.replace(day=1) - relativedelta(months=1)).month
last_day_num = (current_date.replace(day=1) - relativedelta(days=1)).day
last_month_num = (current_date.replace(day=1) - relativedelta(days=1)).month
start_day_str = str(start_day_num)
start_month_str = f"{start_month_num:02}"
last_day_str = str(last_day_num)
last_month_str = f"{last_month_num:02}"

# 1. Открыть https://www.moex.com;
option = webdriver.ChromeOptions()
option.add_argument("--start-maximized")
driver = webdriver.Chrome(options = option)
driver.get('https://www.moex.com')

# 2. Перейти по следующим элементам: Меню -> Срочный рынок -> Индикативные курсы;
driver.find_element(By.CSS_SELECTOR, 'button.header__button[aria-label="Открыть бургер меню"]').click()
time.sleep(3)
driver.find_element(By.LINK_TEXT, 'Срочный рынок').click()
time.sleep(3)
driver.find_element(By.LINK_TEXT, 'Согласен').click()
time.sleep(3)
driver.find_element(By.LINK_TEXT, 'Индикативные курсы').click()
time.sleep(3)

# 3. В выпадающем списке выбрать валюты: USD/RUB - Доллар США к российскому рублю;
wait = WebDriverWait(driver, 10)
dropdown = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '.ui-select__activator'))
)
dropdown.click()
wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'ui-dropdown-option')))
options = driver.find_elements(By.CLASS_NAME, 'ui-dropdown-option')

for option in options:
    if "USD/RUB - Доллар США к российскому рублю" in option.text:
        driver.execute_script("arguments[0].scrollIntoView(true);", option)
        option.click()
        break

# 4. Сформировать данные за предыдущий месяц;
 # Выбор 1 числа предыдущего месяца
callendarClick = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="fromDate"].date_btn'))
)
callendarClick.click()

wait = WebDriverWait(driver, 10)
startMonthClick = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '.ui-group-item.ui-select.ui-calendar__select'))
)
startMonthClick.click()

options = driver.find_elements(By.CLASS_NAME, 'ui-dropdown-option')

for option in options:
    if start_month_str in option.text:
        option.click()
        break

startDayClick = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '.ui-calendar__cell.-day'))
)
startDayClick.click()

time.sleep(5)

 # Выбор последнего числа предыдущего месяца
callendarClick = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="tillDate"].date_btn'))
)
callendarClick.click()

wait = WebDriverWait(driver, 10)
lastMonthClick = wait.until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="x6yzHnS3PmDg26JWSJJHQUtgRdYhwntYBHnpGe6f7KXJI0S58gvlQbsdGrE0XLam"]/div[7]/div[1]/div[1]'))
)
lastMonthClick.click()

options = driver.find_elements(By.CLASS_NAME, 'ui-dropdown-option')

for option in options:
    if last_month_str in option.text:
        option.click()
        break

days = WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.ui-calendar__cell.-day'))
)

for day in days:
    if day.text == last_day_str:
        day.click()
        break

 # Клик по "Показать" и ожидание
clickButtonShow = wait.until(
    EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'dateTimeRowButton')]//span[text()=' Показать ']"))
)
clickButtonShow.click()

time.sleep(5)

# 5. Копирование данных в Excel
wait = WebDriverWait(driver, 10)
table = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'ui-table__container')))

rows = table.find_elements(By.TAG_NAME, 'tr')

data = []
for row in rows[2:]:
    cells = row.find_elements(By.TAG_NAME, 'td')

    if len(cells) == 5:
        date = cells[0].text
        value_usd = cells[3].text
        time_usd = cells[4].text

        data.append([date, value_usd, time_usd])

columns = ['Дата USD/RUB', 'Курс USD/RUB', 'Время USD/RUB']
df = pd.DataFrame(data, columns=columns)

df.to_excel('report.xlsx', index=False)

# 6. В выпадающем списке выбрать валюты: JPY/RUB - Японская йена к российскому рублю;
wait = WebDriverWait(driver, 10)
dropdown = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '.ui-select__activator'))
)
dropdown.click()
wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'ui-dropdown-option')))
options = driver.find_elements(By.CLASS_NAME, 'ui-dropdown-option')

for option in options:
    if "JPY/RUB - Японская йена к российскому рублю" in option.text:
        driver.execute_script("arguments[0].scrollIntoView(true);", option)
        option.click()
        break

#Сформировать данные за предыдущий месяц;
 # Выбор 1 числа предыдущего месяца
callendarClick = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="fromDate"].date_btn'))
)
callendarClick.click()

wait = WebDriverWait(driver, 10)
startMonthClick = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '.ui-group-item.ui-select.ui-calendar__select'))
)
startMonthClick.click()

options = driver.find_elements(By.CLASS_NAME, 'ui-dropdown-option')

for option in options:
    if start_month_str in option.text:
        option.click()
        break

startDayClick = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '.ui-calendar__cell.-day'))
)
startDayClick.click()

 # Выбор последнего числа предыдущего месяца
callendarClick = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="tillDate"].date_btn'))
)
callendarClick.click()

wait = WebDriverWait(driver, 10)
lastMonthClick = wait.until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="x6yzHnS3PmDg26JWSJJHQUtgRdYhwntYBHnpGe6f7KXJI0S58gvlQbsdGrE0XLam"]/div[7]/div[1]/div[1]'))
)
lastMonthClick.click()

options = driver.find_elements(By.CLASS_NAME, 'ui-dropdown-option')

for option in options:
    if last_month_str in option.text:
        option.click()
        break

days = WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.ui-calendar__cell.-day'))
)

for day in days:
    if day.text == last_day_str:
        day.click()
        break

 # Клик по "Показать" и ожидание
clickButtonShow = wait.until(
    EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'dateTimeRowButton')]//span[text()=' Показать ']"))
)
clickButtonShow.click()

file_path = 'report.xlsx'

# 7. Копирование данных в Excel
wait = WebDriverWait(driver, 10)
table = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'ui-table__container')))

rows = table.find_elements(By.TAG_NAME, 'tr')

data = []
for row in rows[2:]:
    cells = row.find_elements(By.TAG_NAME, 'td')

    if len(cells) == 5:
        date = cells[0].text
        value_jpy = cells[3].text
        time_jpy = cells[4].text

        data.append([date, value_jpy, time_jpy])

driver.quit()

wb = load_workbook(file_path)
ws = wb.active

columns = ws.max_column + 1
while ws.cell(row=1, column=columns).value is not None:
    columns += 1

ws.cell(row=1, column=columns, value='Дата JPY/RUB')
ws.cell(row=1, column=columns + 1, value='Курс JPY/RUB')
ws.cell(row=1, column=columns + 2, value='Время JPY/RUB')

start_row = 2

for i, (date, value, time) in enumerate(data, start=start_row):
    ws.cell(row=i, column=columns, value=date)
    ws.cell(row=i, column=columns + 1, value=value_jpy)
    ws.cell(row=i, column=columns + 2, value=time_jpy)

wb.save(file_path)

# 8. Для каждой строки полученного файла поделить курс USD/RUB на JPY/RUB, полученное значение записать в ячейку (G) Результат;

wb = load_workbook(file_path)
ws = wb.active
ws['G1'] = 'Результат'

for row in range(2, ws.max_row + 1):
    value_b = ws.cell(row=row, column=2).value
    value_e = ws.cell(row=row, column=5).value

    if value_b is not None and value_e is not None and value_e != 0:
        result = float(value_b) / float(value_e)
        ws.cell(row=row, column=7, value=result)

# 9. Выровнять – автоширина;
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# 10. Формат чисел – финансовый;
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
    for cell in row:
        cell.number_format = '#,##0.00 "₽";[Red]#,##0.00 "₽"'

# 11. Проверить, что автосумма в Excel распознаёт ячейки как числовой формат;
for cell in ws['B']:
    if cell.row != 1:
        cell.number_format = '#,##0.00'

for cell in ws['E']:
    if cell.row != 1:
        cell.number_format = '#,##0.00'

wb.save(file_path)

# 12. Направить итоговый файл отчета себе на почту;
def get_row_declension(n):
    if 11 <= n % 100 <= 19:
        return f"{n} строк"
    elif n % 10 == 1:
        return f"{n} строка"
    elif 2 <= n % 10 <= 4:
        return f"{n} строки"
    else:
        return f"{n} строк"

def send_email_report(sender_email, sender_password, recipient_email, file_path, row_count):
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = "Итоговый отчет"

    body = f"Отчет с количеством строк: {get_row_declension(row_count)}."
    msg.attach(MIMEText(body, 'plain'))

    attachment = open(file_path, "rb")
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(file_path)}")
    msg.attach(part)

    server = smtplib.SMTP('smtp.mail.ru', 465)
    server.starttls()
    server.login(sender_email, sender_password)
    text = msg.as_string()
    server.sendmail(sender_email, recipient_email, text)
    server.quit()

# 13. В письме указать количество строк в Excel в правильном склонении.
wb = load_workbook(file_path)
ws = wb.active

row_count = ws.max_row - 1
wb.save(file_path)

sender_email = "s0ftcore@mail.ru"
sender_password = ""
recipient_email = "softcore@vk.com"

send_email_report(sender_email, sender_password, recipient_email, file_path, row_count)
print(f"Отчет отправлен на {recipient_email}. Количество строк: {get_row_declension(row_count)}.")